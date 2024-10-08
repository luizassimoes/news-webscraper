import re
import time
import logging
import urllib.parse
from datetime import datetime, timedelta
from dateutil import parser
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from RPA.core.webdriver import start
from RPA.Robocorp.WorkItems import WorkItems

from topics_dict import topics_dict  # Contains the URL code for each Topic: invoked within the select_topic()


class NewsWebScraper:

    def __init__(self):
        self.driver = None
        self.logger = logging.getLogger(__name__)

    def set_chrome_options(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-web-security')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument("--log-level=3")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        return options

    def set_webdriver(self, browser='Chrome'):
        options = self.set_chrome_options()
        try:
            self.driver = start(browser, options=options)
            self.logger.info('WebDriver started successfully.')
        except Exception as e:
            self.logger.error(f'ERR0R set_webdriver() | Failed to start WebDriver: {e}')

    def open_url(self, url: str):
        if self.driver:
            try:
                self.driver.get(url)
                self.logger.info(f'Opened URL: {url}')
            except Exception as e:
                self.logger.error(f'ERROR open_url() | Failed to open URL {url}: {e}')
        else:
            self.logger.error('ERR0R open_url() | WebDriver not initialized. You must call set_webdriver() first.')

    def search(self, url: str, search_query: str):
        """Perform a search on the given URL using the provided query string."""
        try:
            search_url = url + 'search?q=' + urllib.parse.quote(search_query)
            self.logger.info(f'Search query submitted: {search_query}.')
            self.open_url(search_url)
        except Exception as e:
            self.logger.error(f'ERROR search() | Could not find element: {e}')

    def sort_by_newest(self):
        current_url = self.driver.current_url
        next_url = current_url + '&s=1'  # Parameter to sort by Newest: s=1
        self.logger.info(f'Query sorted by Newest.')    
        self.open_url(next_url)

    def select_topic(self, topic):
        topic = topic.title()
        if topic in topics_dict.keys():
            current_url = self.driver.current_url
            url_0, url_1 = current_url.rsplit('&', 1)
            url_topic = url_0 + '&' + topics_dict[topic] + '&' + url_1  # Topic goes between search and sort by
            self.logger.info(f"Selected topic: {topic}.")
            self.open_url(url_topic)
        else:
            self.logger.warning(f'The topic {topic} does not exist. No topic was selected.')
        
    def get_element_list(self, element_selector, src=False):
        """
        Gets the elements from the web page. The src parameter allows the function to treat img tags.
        The function captures all the news blocks and checks if they have the according element in it. 
        If not, it registers an empty string so the order of the information is correctly organized in
        the final Excel file.        
        """
        wait = WebDriverWait(self.driver, 10)
        retries = 2
        element_name = element_selector.split('-')[-1].split(' ')[0].title()  # Gets the element name to show in the log
        for attempt in range(retries):  # Implement retries in case the page takes long to load
            try: 
                elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, element_selector)))
                news_cards = self.driver.find_elements(By.CLASS_NAME, 'promo-wrapper')
                
                element_idx = 0
                element_list = []
                for card in news_cards:
                    has_element = card.find_elements(By.CSS_SELECTOR, element_selector)
                    if has_element:
                        if not src:
                            element_list.append(elements[element_idx].text)
                        else:
                            element_list.append(elements[element_idx].get_attribute('src'))
                        element_idx += 1
                    else:
                        element_list.append('')
                self.logger.info(f'{element_name}s gotten on attempt {attempt+1}.')
                return element_list
            except Exception as e:
                if attempt == retries - 1:
                    self.logger.warning(f'Element not found: {element_name}.')
                else:
                    self.logger.warning(f'Could not get {element_name} on attempt {attempt+1}. Retrying...')
    
    def download_pics(self, pic_urls):
        filenames = []
        downloads = 0
        for i, url in enumerate(pic_urls):
            try:
                if downloads < 47:  # Total amount of files in Robocorp is 50 = Excel File + console.txt + env.yaml + 47 images
                    self.driver.get(url)
                    width = 840
                    height = 560
                    self.driver.set_window_size(width, height)

                    filename = f'image_{str(i+1).zfill(2)}.png'
                    self.driver.save_screenshot(f'./output/{filename}')
                    self.logger.info(f'Downloaded {filename}.')
                    filenames.append(filename)
                    downloads += 1
                else:
                    filenames.append('')
                    self.logger.warning(f'LIMIT EXCEEDED: Cannot upload more files to Robocorp.')
            except Exception as e:
                self.logger.warning(f'No image for news in row {i+2}.')  # Two more to reference the header and skip 0
                filenames.append('')
        return filenames

    def count_search_query(self, query, titles, descriptions):
        query_count = []
        for i in range(len(titles)):
            title_count = titles[i].count(query)
            description_count = descriptions[i].count(query)
            query_count.append(title_count + description_count)
        self.logger.info(f'Search query in Title counted.')
        return query_count
    
    def title_contains_money(self, titles):
        money_patterns = [
            r'\$\d{1,3}(,\d{3})*(\.\d{2})?',      # Pattern for $11.1 or $111,111.11
            r'\b\d+(\.\d{2})?\s?(dollars|USD)\b'  # Pattern for 11 dollars or 11 USD
        ]
        money_in_query = []
        for title in titles:
            money = False
            for pattern in money_patterns:
                if re.search(pattern, title, re.IGNORECASE):
                    money = True
                    break
            money_in_query.append(money)
        self.logger.info(f'Contains money analyzed.')
        return money_in_query

    def next_page(self):
        """Uses the 'p' element to navagate through pages via the URL."""
        current_url = self.driver.current_url
        url_preffix, url_suffix = current_url.rsplit('&', 1)  # The 'p' element is always the last one in the URL
        if url_suffix.startswith('p='):
            next_url = url_preffix + '&p=' +  str(int(url_suffix.rsplit('=', 1)[-1])+1)
        else:  # No 'p' element in the URL means the current page is page 1
            next_url = current_url + '&p=2'
        self.logger.info(f"Next page.")
        self.open_url(next_url)

    def parse_date(self, date_str):
        try:
            date = parser.parse(str(date_str))
            return date
        except:  # Date format is hours or minutes
            match_hour = re.match(r'(\d+)\s+hour[s]?\s+ago', str(date_str))      # Pattern for X hours ago
            match_minute = re.match(r'(\d+)\s+minute[s]?\s+ago', str(date_str))  # Pattern for X minutes ago
            if match_hour:
                hours_ago = int(match_hour.group(1))
                date = datetime.now() - timedelta(hours=hours_ago)
                return date
            elif match_minute:
                minutes_ago = int(match_minute.group(1))
                date = datetime.now() - timedelta(minutes=minutes_ago)
                return date
            else:
                self.logger.error(f'ERROR parse_date() | Date "{date_str}" is invalid. Could not get date.')

    def get_news(self, n):  # n = number of months
        titles = []
        descriptions = []
        dates = []
        pic_urls = []
        filenames = []
        out_of_date =  False

        try: 
            n = int(n)
        except:
            self.logger.warning(f'The variable MONTHS has to be a number, not {n}. Using 1 month instead.')
            n = 1

        if n < 0:
            self.logger.warning(f'The number of months can not be negative: {n}. Using {abs(n)} instead.')
        n = 1 if n == 0 else abs(n)

        tomorrow = datetime.now() + timedelta(days=1)
        try:
            n_months_ago = tomorrow - relativedelta(months=n)
            self.logger.info(f'Getting news until {n} month(s) ago.')
        except:
            n_months_ago = tomorrow - relativedelta(months=1)
            invalid_year = tomorrow.year - int(n/12)
            self.logger.warning(f'Year {invalid_year} is out of range. Using 1 month instead.')

        
        while True:
            aux_titles = self.get_element_list('h3.promo-title a')
            if aux_titles is None:
                break

            aux_descriptions = self.get_element_list('p.promo-description')
            aux_dates = self.get_element_list('p.promo-timestamp')
            get_pic_urls = self.get_element_list('picture img.image', src=True)
            aux_pic_urls = get_pic_urls if get_pic_urls is not None else []

            for i, date_str in enumerate(aux_dates):
                date_obj = self.parse_date(date_str)

                aux_dates[i] = date_obj
                if date_obj < n_months_ago:
                    out_of_date = True
                    break
        
            if out_of_date:  # Gets only the list items that are in the specified period 
                titles.extend(aux_titles[:i].copy())
                descriptions.extend(aux_descriptions[:i].copy())
                dates.extend(aux_dates[:i].copy())
                pic_urls.extend(aux_pic_urls[:i].copy())
                break
            else:  # Gets the news from the current page and goes to the next page
                titles.extend(aux_titles.copy())
                descriptions.extend(aux_descriptions.copy())
                dates.extend(aux_dates.copy())
                pic_urls.extend(aux_pic_urls.copy())
                self.next_page()     

                current_url = self.driver.current_url
                url_suffix = current_url.rsplit('&', 1)[-1]
                if not url_suffix.startswith('p='):  # It means the amount of pages is over
                    break

        filenames = self.download_pics(pic_urls)

        return titles, descriptions, dates, filenames


    def close_all(self):
        if self.driver:
            try:
                self.driver.quit()
                self.logger.info('WebDriver closed successfully.')
            except Exception as e:
                self.logger.error(f'ERROR close_all() | Failed to close WebDriver: {e}')
        else:
            self.logger.error('ERROR close_all() | WebDriver not initialized.')

    def to_excel(self, data):
        wb = Workbook()
        sheet = wb.active

        headers = ['Title', 'Description', 'Date', 'Filename', 'Count Query in Title and Description', 'Contains Money in Title']
        for i_col, header in enumerate(headers):
            cell_header = sheet.cell(row=1, column=i_col+1)
            cell_header.value = header
            cell_header.font = Font(bold=True, size=12)
            cell_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if data[i_col] is not None:
                for i_row, val in enumerate(data[i_col]):
                    cell_content = sheet.cell(row=i_row+2, column=i_col+1)
                    cell_content.value = val
                    sheet.row_dimensions[i_row+2].height = 60
                    if i_col > 1:  # Horizontal aligment not for columns Title and Description
                        cell_content.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell_content.alignment = Alignment(vertical='center', wrap_text=True)
            else:
                self.logger.error(f'ERROR to_excel() | {header} list is None.')

        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F']:
            sheet.column_dimensions[col].width = 25
        sheet.row_dimensions[1].height = 35

        sheet.title = 'NEWS'
        self.logger.info('Excel done.')
        return wb


def main():
    url = 'https://www.latimes.com/'

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    work_items = WorkItems()
    work_items.get_input_work_item()  # Gets the variables from the Robocloud Workspace
    query = work_items.get_work_item_variable("SEARCH_QUERY", default='teste')
    query = str(query)
    topic = work_items.get_work_item_variable("TOPIC", default='Business')
    topic = str(topic).strip()
    n_months = work_items.get_work_item_variable("MONTHS", default=0)

    news_scraper = NewsWebScraper()
    news_scraper.logger.info('-'*60)
    news_scraper.logger.info('Variables imported. Process is starting.')

    news_scraper.set_webdriver()
    news_scraper.search(url, query)
    news_scraper.sort_by_newest()
    time.sleep(1)
    news_scraper.select_topic(topic)
    time.sleep(1)

    titles, descriptions, dates, filenames = news_scraper.get_news(n_months)
    if titles:
        count_query = news_scraper.count_search_query(query, titles, descriptions)
        contains_money = news_scraper.title_contains_money(titles)

        news_data = [titles, descriptions, dates, filenames, count_query, contains_money]
        
        wb = news_scraper.to_excel(news_data)
        wb.save('./output/News.xlsx')
    else:
        news_scraper.logger.error(f'Your search returned no News: {query}, topic: {topic}, months: {n_months}.')

    news_scraper.close_all()

    news_scraper.logger.info('-'*60)


if __name__ == '__main__':
    main()
