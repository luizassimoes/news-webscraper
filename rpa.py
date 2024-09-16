import os
import re
import time
import logging
import requests
from datetime import datetime, timedelta
from dateutil import parser
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from RPA.core.webdriver import start


class NewsWebScraper:

    def __init__(self):
        self.driver = None
        self.logger = logging.getLogger(__name__)

    def set_chrome_options(self):
        options = webdriver.ChromeOptions()
        # options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-web-security')
        options.add_argument('--start-maximized')
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        return options

    def set_webdriver(self, browser='Chrome'):
        options = self.set_chrome_options()
        try:
            self.driver = start(browser, options=options)
            self.logger.info('WebDriver started successfully.')
        except Exception as e:
            self.logger.error(f'ERR0R set_webdriver() | Failed to start WebDriver: {e}')

    def open_url(self, url: str):
        if self.driver:
            try:
                self.driver.get(url)
                self.logger.info(f'Opened URL: {url}')
            except Exception as e:
                self.logger.error(f'ERROR open_url() | Failed to open URL {url}: {e}')
        else:
            self.logger.error('ERR0R open_url() | WebDriver not initialized. You must call set_webdriver() first.')

    def search(self, search_query: str):
        try:
            search_button = self.driver.find_element(By.XPATH, '/html/body/ps-header/header/div[2]/button')
            search_button.click()
            self.logger.info(f'Search button clicked.')
            search_field = self.driver.find_element(By.XPATH, '/html/body/ps-header/header/div[2]/div[2]/form/label/input')
            search_field.send_keys(search_query)
            search_field.submit()
            self.logger.info(f'Search query submitted.')
        except Exception as e:
            self.logger.error(f'ERROR search() | Could not find element: {e}')

    def sort_by_newest(self):
        current_url = self.driver.current_url
        next_url = current_url + '&s=1'
        self.logger.info(f'Query sorted by Newest.')    
        self.open_url(next_url)

    def select_topic(self, topic):
        wait = WebDriverWait(self.driver, 10)
        try:
            see_all_btn = wait.until(EC.presence_of_element_located((By.XPATH, '//button[span[@class="see-all-text" and text()="See All"]]')))
            see_all_btn.click()
            checkbox = wait.until(EC.presence_of_element_located((By.XPATH, f'//label[span[text()="{topic}"]]/input')))
            checkbox.click()
            self.logger.info(f"Selected the topic: {topic}.")
        except Exception as e:
            self.logger.error(f"ERROR select_topic() | Could not find the topic: '{topic}'.")
        
    def get_element_list(self, element_selector, src=False):
        wait = WebDriverWait(self.driver, 10)
        retries = 3
        element_name = element_selector.split('-')[-1].split(' ')[0].title()
        for attempt in range(retries):
            try: 
                elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, element_selector)))
                self.logger.info(f'{element_name}s gotten on attempt {attempt+1}.')
                if not src:
                    return [element.text for element in elements]
                else:
                    return [element.get_attribute('src') for element in elements]
            except Exception as e:
                self.logger.warning(f'Could not get {element_name} on attempt {attempt+1}. Retrying...')
                if attempt == retries - 1:
                    self.logger.error(f'ERROR get_element_list(): {element_name} | Error: {e}')
    
    def download_pics(self, pic_urls):
        filenames = []
        for i, url in enumerate(pic_urls):
            try:
                response = requests.get(url)
                response.raise_for_status()
                filename = f'image_{i+1}.jpeg'
                filepath = f'./outputs/{filename}'
                with open(filepath, 'wb') as file:
                    file.write(response.content)
                filenames.append(filename)
                self.logger.info(f'Downloaded {filename}.')
            except requests.exceptions.RequestException as e:
                self.logger.error(f'Failed to download {url}: {e}')
        return filenames

    def count_search_query(self, query, titles, descriptions):
        query_count = []
        for i in range(len(titles)):
            title_count = titles[i].count(query)
            description_count = descriptions[i].count(query)
            query_count.append(title_count + description_count)
        self.logger.info(f'Search query in Title counted.')
        return query_count
    
    def title_contains_money(self, titles):
        money_patterns = [
            r'\$\d{1,3}(,\d{3})*(\.\d{2})?',      # Pattern for $11.1 or $111,111.11
            r'\b\d+(\.\d{2})?\s?(dollars|USD)\b'  # Pattern for 11 dollars or 11 USD
        ]
        money_in_query = []
        for title in titles:
            money = False
            for pattern in money_patterns:
                if re.search(pattern, title, re.IGNORECASE):
                    money = True
                    break
            money_in_query.append(money)
        self.logger.info(f'Contains money analyzed.')
        return money_in_query

    def next_page(self):
        wait = WebDriverWait(self.driver, 100)
        try:
            next_btn = wait.until(EC.presence_of_element_located((By.XPATH, '//a[span[text()="Next"]]')))
            next_btn.click()
            self.logger.info(f"Next page.")
            time.sleep(2)
        except Exception as e:
            self.logger.error(f'ERROR next_page() | Could not change pages.')

    def parse_date(self, date_str):
        try:
            date = parser.parse(str(date_str))
            return date
        except ValueError:
            self.logger.error(f'ERROR parse_date() | Date "{date_str}" not in format Month Day, Year.')

        match = re.match(r'(\d+)\s+hour[s]?\s+ago', date_str)
        if match:
            hours_ago = int(match.group(1))
            date = datetime.now() - timedelta(hours=hours_ago)
            return date
        else:
            self.logger.error(f'ERROR parse_date() | Date "{date_str}" not in format X hours ago. Could not get date.')

    def get_news(self, n):
        out_of_date =  False

        n = 1 if n == 0 else n
        tomorrow = datetime.now() + timedelta(days=1)
        n_months_ago = tomorrow - relativedelta(months=n)
        
        while True:
            titles = self.get_element_list('h3.promo-title a')
            descriptions = self.get_element_list('p.promo-description')
            dates = self.get_element_list('p.promo-timestamp')
            pic_urls = self.get_element_list('picture img.image', src=True)

            for i, date_str in enumerate(dates):
                date_obj = self.parse_date(date_str)
                print(date_obj)

                if date_obj is not None:
                    dates[i] = date_obj
                    if date_obj < n_months_ago:
                        out_of_date = True
                        break
        
            if out_of_date:
                break
            else:
                self.next_page()     

        titles = titles[0:i].copy()
        descriptions = descriptions[0:i].copy()
        dates = dates[0:i].copy()
        pic_urls = pic_urls[0:i].copy()

        filenames = self.download_pics(pic_urls)

        return titles, descriptions, dates, filenames

    def close_all(self):
        if self.driver:
            try:
                self.driver.quit()
                self.logger.info('WebDriver closed successfully.')
            except Exception as e:
                self.logger.error(f'ERROR close_all() | Failed to close WebDriver: {e}')
        else:
            self.logger.error('ERROR close_all() | WebDriver not initialized.')

    def to_excel(self, data):
        wb = Workbook()
        sheet = wb.active

        headers = ['Title', 'Description', 'Date', 'Filename', 'Count Query in Title', 'Contains Money in Title']
        for i_col, header in enumerate(headers):
            sheet.cell(row=1, column=i_col+1).value = header

            if data[i_col] is not None:
                for i_row, val in enumerate(data[i_col]):
                    sheet.cell(row=i_row+2, column=i_col+1).value = val
            else:
                self.logger.error(f'ERROR to_excel() | {header} list is None.')
        sheet.title = 'NEWS'
        self.logger.info('Excel done.')
        return wb

def main():
    url = 'https://www.latimes.com/'
    query = 'euro'
    topic = 'Books'
    n_months = 2

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    news_scraper = NewsWebScraper()
    news_scraper.set_webdriver()
    news_scraper.open_url(url)
    news_scraper.search(query)
    news_scraper.sort_by_newest()
    time.sleep(1)
    news_scraper.select_topic(topic)
    time.sleep(1)

    titles, descriptions, dates, filenames = news_scraper.get_news(n_months)
    count_query = news_scraper.count_search_query(query, titles, descriptions)
    contains_money = news_scraper.title_contains_money(titles)

    news_data = [titles, descriptions, dates, filenames, count_query, contains_money]
    
    wb = news_scraper.to_excel(news_data)
    wb.save('./outputs/News.xlsx')

    time.sleep(3)

    news_scraper.close_all()


if __name__ == '__main__':
    main()
